from pathlib import Path
from io import BytesIO
import importlib
import os
import re
import sqlite3
import subprocess
from datetime import datetime

from flask import Flask, jsonify, redirect, render_template_string, request, send_file, url_for
import pandas as pd

app = Flask(__name__)
excel_path = Path(__file__).with_name("auto_lijst.xlsx")
db_path = Path(__file__).with_name("auto_lijst.db")
static_path = Path(__file__).with_name("static")
table_name = "autos"
default_columns = ["Merk", "Type", "Bouwjaar", "Prijs", "Categorie"]
filter_field_names = [
    "filter_merk",
    "filter_type",
    "filter_bouwjaar_min",
    "filter_bouwjaar_max",
    "filter_prijs_min",
    "filter_prijs_max",
    "filter_categorie",
]


def get_app_version() -> str:
    env_version = os.getenv("RENDER_GIT_COMMIT") or os.getenv("GIT_COMMIT")
    if env_version:
        return env_version[:7]

    try:
        output = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], stderr=subprocess.DEVNULL)
        return output.decode("utf-8").strip()
    except Exception:
        return "unknown"


def get_header_image_url() -> str | None:
    candidates = [
        "porsche911cabrio.jpg",
    ]

    for filename in candidates:
        if (static_path / filename).exists():
            return url_for("static", filename=filename)

    return None


def ensure_category_column(df: pd.DataFrame) -> tuple[pd.DataFrame, bool]:
    changed = False

    if "Categorie" not in df.columns:
        df["Categorie"] = ""
        changed = True

    if "Categorie" in df.columns and df["Categorie"].dtype != object:
        df["Categorie"] = df["Categorie"].astype("object")
        changed = True

    ordered_columns = [column for column in df.columns if column != "Categorie"] + ["Categorie"]
    if list(df.columns) != ordered_columns:
        df = df[ordered_columns]
        changed = True

    return df, changed


def parse_bouwjaar(value) -> int | None:
    if pd.isna(value):
        return None

    if isinstance(value, (int, float)):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    text = str(value).strip()
    match = re.search(r"(19\d{2}|20\d{2})", text)
    if not match:
        return None

    try:
        return int(match.group(1))
    except ValueError:
        return None


def determine_category_from_bouwjaar(bouwjaar_value) -> str:
    year = parse_bouwjaar(bouwjaar_value)
    if year is None:
        return ""
    return "Klassieker" if year < 1990 else "Youngtimer"


def apply_category_rules(df: pd.DataFrame) -> tuple[pd.DataFrame, bool]:
    if "Bouwjaar" not in df.columns or "Categorie" not in df.columns:
        return df, False

    changed = False
    for row_index in df.index:
        new_category = determine_category_from_bouwjaar(df.at[row_index, "Bouwjaar"])
        current_category = "" if pd.isna(df.at[row_index, "Categorie"]) else str(df.at[row_index, "Categorie"])
        if current_category != new_category:
            df.at[row_index, "Categorie"] = new_category
            changed = True

    return df, changed


def render_file_locked_message(action_label: str):
    return (
        render_template_string(
            """
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 0;
                    background: #f6f8fb;
                    color: #1f2937;
                }
                .container {
                    max-width: 760px;
                    margin: 32px auto;
                    background: white;
                    border: 1px solid #e5e7eb;
                    border-radius: 10px;
                    padding: 24px;
                }
                .warn {
                    background: #fff7ed;
                    border: 1px solid #fdba74;
                    border-radius: 8px;
                    padding: 12px;
                    margin-top: 10px;
                }
                a {
                    display: inline-block;
                    margin-top: 14px;
                    text-decoration: none;
                    color: #1d4ed8;
                }
            </style>
            <div class="container">
                <h1>Opslaan lukt niet</h1>
                <div class="warn">
                    De wijziging ({{ action_label }}) kon niet worden opgeslagen.
                </div>
                <p>Probeer het opnieuw. Als het probleem blijft, herstart de app.</p>
                <a href="{{ url_for('home') }}">← Terug naar overzicht</a>
            </div>
            """,
            action_label=action_label,
        ),
        409,
    )


def table_exists(connection: sqlite3.Connection, name: str) -> bool:
    cursor = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (name,),
    )
    return cursor.fetchone() is not None


def load_dataframe() -> pd.DataFrame:
    if not db_path.exists() and excel_path.exists():
        try:
            migrated = pd.read_excel(excel_path)
            migrated, _ = ensure_category_column(migrated)
            migrated, _ = apply_category_rules(migrated)
            save_dataframe(migrated)
        except Exception:
            pass

    if not db_path.exists():
        return pd.DataFrame(columns=default_columns)

    with sqlite3.connect(db_path) as connection:
        if not table_exists(connection, table_name):
            return pd.DataFrame(columns=default_columns)

        query = f'SELECT id, "Merk", "Type", "Bouwjaar", "Prijs", "Categorie" FROM "{table_name}" ORDER BY id'
        df = pd.read_sql_query(query, connection)

    return df


def save_dataframe(df: pd.DataFrame) -> None:
    df, _ = ensure_category_column(df)
    df, _ = apply_category_rules(df)

    if "id" not in df.columns:
        df["id"] = pd.NA

    for column in default_columns:
        if column not in df.columns:
            df[column] = ""

    df = df[["id", *default_columns]].copy()
    df = df.fillna("")
    for column in default_columns:
        df[column] = df[column].astype(str)

    with sqlite3.connect(db_path) as connection:
        connection.execute(
            f'''
            CREATE TABLE IF NOT EXISTS "{table_name}" (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                "Merk" TEXT,
                "Type" TEXT,
                "Bouwjaar" TEXT,
                "Prijs" TEXT,
                "Categorie" TEXT
            )
            '''
        )
        connection.execute(f'DELETE FROM "{table_name}"')

        insert_with_id_sql = f'INSERT INTO "{table_name}" (id, "Merk", "Type", "Bouwjaar", "Prijs", "Categorie") VALUES (?, ?, ?, ?, ?, ?)'
        insert_without_id_sql = f'INSERT INTO "{table_name}" ("Merk", "Type", "Bouwjaar", "Prijs", "Categorie") VALUES (?, ?, ?, ?, ?)'

        for _, row in df.iterrows():
            parsed_id = parse_row_id(row.get("id"))
            values = tuple(row[column] for column in default_columns)
            if parsed_id is not None:
                connection.execute(insert_with_id_sql, (parsed_id, *values))
            else:
                connection.execute(insert_without_id_sql, values)

        connection.commit()


def format_eur_value(value) -> str:
    if pd.isna(value):
        return ""

    if isinstance(value, str):
        cleaned = (
            value.strip()
            .replace("€", "")
            .replace(" ", "")
            .replace(".-", "")
            .replace(",-", "")
            .replace(".", "")
            .replace(",", ".")
        )
    else:
        cleaned = str(value)

    try:
        number = float(cleaned)
    except ValueError:
        return str(value)

    if number.is_integer():
        thousands = f"{int(number):,}".replace(",", ".")
        return f"€ {thousands},-"

    formatted = f"{number:,.2f}"
    formatted = formatted.replace(",", "_").replace(".", ",").replace("_", ".")
    return f"€ {formatted}"


def to_display_value(value, column_name: str | None = None, format_price: bool = False) -> str:
    if pd.isna(value):
        return ""

    if format_price and column_name and column_name.strip().lower() == "prijs":
        return format_eur_value(value)

    return str(value)


def parse_positive_int(value: str | None, default: int) -> int:
    try:
        parsed = int(value) if value is not None else default
    except ValueError:
        return default
    return parsed if parsed > 0 else default


def parse_optional_int(value: str | None) -> int | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        return int(text)
    except ValueError:
        return None


def parse_price_number(value) -> float | None:
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    cleaned = (
        text.replace("€", "")
        .replace(" ", "")
        .replace(".-", "")
        .replace(",-", "")
    )

    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "")
        cleaned = cleaned.replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")

    try:
        return float(cleaned)
    except ValueError:
        return None


def get_filter_values_from_request(values) -> dict[str, str]:
    result: dict[str, str] = {}
    for field in filter_field_names:
        result[field] = str(values.get(field, "")).strip()
    return result


def get_non_empty_filter_params(filter_values: dict[str, str]) -> dict[str, str]:
    return {key: value for key, value in filter_values.items() if value != ""}


def get_sort_values_from_request(values) -> tuple[str, str]:
    allowed_sort_fields = {"id", "Merk", "Type", "Bouwjaar", "Prijs", "Categorie"}
    sort_by = str(values.get("sort_by", "id")).strip()
    sort_dir = str(values.get("sort_dir", "asc")).strip().lower()

    if sort_by not in allowed_sort_fields:
        sort_by = "id"
    if sort_dir not in {"asc", "desc"}:
        sort_dir = "asc"

    return sort_by, sort_dir


def apply_sorting(df: pd.DataFrame, sort_by: str, sort_dir: str) -> pd.DataFrame:
    if df.empty or sort_by not in df.columns:
        return df

    ascending = sort_dir == "asc"
    sorted_df = df.copy()

    if sort_by == "Bouwjaar":
        key_series = sorted_df["Bouwjaar"].apply(parse_bouwjaar)
        fill_value = float("inf") if ascending else float("-inf")
        sorted_df["_sort_key"] = key_series.fillna(fill_value)
    elif sort_by == "Prijs":
        key_series = sorted_df["Prijs"].apply(parse_price_number)
        fill_value = float("inf") if ascending else float("-inf")
        sorted_df["_sort_key"] = key_series.fillna(fill_value)
    elif sort_by == "id":
        key_series = sorted_df["id"].apply(parse_row_id)
        fill_value = float("inf") if ascending else float("-inf")
        sorted_df["_sort_key"] = key_series.fillna(fill_value)
    else:
        sorted_df["_sort_key"] = sorted_df[sort_by].astype(str).str.strip().str.lower()

    sorted_df = sorted_df.sort_values("_sort_key", ascending=ascending, kind="mergesort")
    return sorted_df.drop(columns=["_sort_key"])


def parse_row_id(value) -> int | None:
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        parsed = int(float(text))
    except ValueError:
        return None

    return parsed if parsed > 0 else None


def build_filtered_dataframe(df: pd.DataFrame, filter_values: dict[str, str]) -> pd.DataFrame:
    filtered_df = df.copy()

    filter_merk = filter_values["filter_merk"]
    if filter_merk and "Merk" in filtered_df.columns:
        filtered_df = filtered_df[
            filtered_df["Merk"].astype(str).str.contains(filter_merk, case=False, na=False, regex=False)
        ]

    filter_type = filter_values["filter_type"]
    if filter_type and "Type" in filtered_df.columns:
        filtered_df = filtered_df[
            filtered_df["Type"].astype(str).str.contains(filter_type, case=False, na=False, regex=False)
        ]

    bouwjaar_min = parse_optional_int(filter_values["filter_bouwjaar_min"])
    bouwjaar_max = parse_optional_int(filter_values["filter_bouwjaar_max"])
    if "Bouwjaar" in filtered_df.columns and (bouwjaar_min is not None or bouwjaar_max is not None):
        year_series = filtered_df["Bouwjaar"].apply(parse_bouwjaar)
        if bouwjaar_min is not None:
            filtered_df = filtered_df[year_series.notna() & (year_series >= bouwjaar_min)]
            year_series = filtered_df["Bouwjaar"].apply(parse_bouwjaar)
        if bouwjaar_max is not None:
            filtered_df = filtered_df[year_series.notna() & (year_series <= bouwjaar_max)]

    prijs_min = parse_price_number(filter_values["filter_prijs_min"])
    prijs_max = parse_price_number(filter_values["filter_prijs_max"])
    if "Prijs" in filtered_df.columns and (prijs_min is not None or prijs_max is not None):
        price_series = filtered_df["Prijs"].apply(parse_price_number)
        if prijs_min is not None:
            filtered_df = filtered_df[price_series.notna() & (price_series >= prijs_min)]
            price_series = filtered_df["Prijs"].apply(parse_price_number)
        if prijs_max is not None:
            filtered_df = filtered_df[price_series.notna() & (price_series <= prijs_max)]

    filter_categorie = filter_values["filter_categorie"]
    if filter_categorie and "Categorie" in filtered_df.columns:
        filtered_df = filtered_df[
            filtered_df["Categorie"].astype(str).str.strip().str.lower() == filter_categorie.strip().lower()
        ]

    return filtered_df

@app.route("/")
def home():
    df = load_dataframe()
    app_version = get_app_version()
    header_image_url = get_header_image_url()

    if df.empty and len(df.columns) == 0:
        return render_template_string(
            """
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 0;
                    background: #f6f8fb;
                    color: #1f2937;
                }
                .container {
                    max-width: 900px;
                    margin: 32px auto;
                    background: white;
                    border: 1px solid #e5e7eb;
                    border-radius: 10px;
                    padding: 24px;
                }
                h1 {
                    margin-top: 0;
                }
                .version {
                    color: #6b7280;
                    font-size: 0.85rem;
                }
                .page-head {
                    display: flex;
                    align-items: flex-start;
                    gap: 14px;
                }
                .brand-image {
                    width: 220px;
                    max-width: 38vw;
                    height: auto;
                    border-radius: 10px;
                    object-fit: cover;
                    border: 1px solid #e5e7eb;
                }
            </style>

            <div class="container">
                <div class="page-head">
                    {% if header_image_url %}
                        <img src="{{ header_image_url }}" alt="Auto" class="brand-image">
                    {% endif %}
                    <div>
                        <h1>Autolijst</h1>
                        <p class="version">Versie: {{ app_version }}</p>
                    </div>
                </div>
                <p>Er is nog geen data gevonden.</p>
                <p>Plaats eventueel <strong>auto_lijst.xlsx</strong> in dezelfde map als <strong>app.py</strong> voor een eenmalige migratie.</p>
            </div>
            """,
            app_version=app_version,
            header_image_url=header_image_url,
        )

    df, _ = ensure_category_column(df)
    df, _ = apply_category_rules(df)
    columns = [column for column in df.columns if column != "id"]
    editable_columns = [column for column in columns if column != "Categorie"]

    filter_values = get_filter_values_from_request(request.args)
    filtered_df = build_filtered_dataframe(df, filter_values)
    sort_by, sort_dir = get_sort_values_from_request(request.args)
    filtered_df = apply_sorting(filtered_df, sort_by, sort_dir)

    total_rows = len(filtered_df)
    page = parse_positive_int(request.args.get("page"), 1)
    per_page = parse_positive_int(request.args.get("per_page"), 50)
    per_page = min(per_page, 500)

    total_pages = max(1, (total_rows + per_page - 1) // per_page)
    page = min(page, total_pages)

    start = (page - 1) * per_page
    end = min(start + per_page, total_rows)

    records = []
    for row_index in range(start, end):
        stable_id = parse_row_id(filtered_df.iloc[row_index].get("id"))
        row_data = {
            column: to_display_value(
                filtered_df.iloc[row_index][column],
                column_name=column,
                format_price=True,
            )
            for column in columns
        }
        original_index = int(filtered_df.index[row_index])
        records.append(
            {
                "index": original_index,
                "display_id": stable_id if stable_id is not None else row_index + 1 + start,
                "data": row_data,
            }
        )

    non_empty_filter_params = get_non_empty_filter_params(filter_values)
    base_query_params = {
        "per_page": per_page,
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        **non_empty_filter_params,
    }
    prev_page_url = url_for("home", page=page - 1, **base_query_params) if page > 1 else None
    next_page_url = url_for("home", page=page + 1, **base_query_params) if page < total_pages else None
    pdf_export_url = url_for("export_pdf", sort_by=sort_by, sort_dir=sort_dir, **non_empty_filter_params)
    excel_export_url = url_for("export_excel", sort_by=sort_by, sort_dir=sort_dir, **non_empty_filter_params)

    sort_targets = ["id", *columns]
    sort_urls: dict[str, str] = {}
    sort_labels: dict[str, str] = {}
    for target in sort_targets:
        next_dir = "desc" if sort_by == target and sort_dir == "asc" else "asc"
        arrow = ""
        if sort_by == target:
            arrow = " ↑" if sort_dir == "asc" else " ↓"
        sort_labels[target] = f"{target}{arrow}"
        sort_urls[target] = url_for(
            "home",
            page=1,
            per_page=per_page,
            sort_by=target,
            sort_dir=next_dir,
            **non_empty_filter_params,
        )

    return render_template_string(
        """
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 0;
                background: #f6f8fb;
                color: #1f2937;
            }
            .container {
                max-width: 1000px;
                margin: 28px auto;
                background: white;
                border: 1px solid #e5e7eb;
                border-radius: 10px;
                padding: 24px;
            }
            h1 {
                margin-top: 0;
                margin-bottom: 18px;
            }
            .version {
                color: #6b7280;
                font-size: 0.85rem;
                margin-top: -8px;
                margin-bottom: 0;
            }
            .top-layout {
                display: flex;
                align-items: stretch;
                gap: 22px;
                margin-bottom: 24px;
            }
            .brand-panel {
                width: 280px;
                flex: 0 0 280px;
                display: flex;
            }
            .page-head {
                display: flex;
                flex-direction: column;
                justify-content: space-between;
                height: 100%;
                width: 100%;
            }
            .brand-image {
                width: 100%;
                max-width: 250px;
                height: auto;
                margin-top: 12px;
                border-radius: 10px;
                object-fit: cover;
                border: 1px solid #e5e7eb;
            }
            @media (max-width: 1000px) {
                .top-layout {
                    flex-direction: column;
                }
                .brand-panel {
                    width: 100%;
                    flex: 1 1 auto;
                }
                .brand-image {
                    max-width: 380px;
                }
            }
            h2 {
                margin: 0;
                font-size: 1.2rem;
            }
            .data-table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 18px;
                font-size: 0.92rem;
            }
            .data-table th,
            .data-table td {
                border: 1px solid #d1d5db;
                padding: 8px;
                text-align: left;
                vertical-align: top;
            }
            .data-table th {
                background: #f3f4f6;
                color: #111827;
            }
            .sort-link {
                color: inherit;
                text-decoration: none;
                display: inline-block;
            }
            .sort-link:hover {
                text-decoration: underline;
            }
            .field {
                margin-bottom: 10px;
            }
            .add-section {
                border: 2px solid #bfdbfe;
                background: #eff6ff;
                border-radius: 10px;
                padding: 16px;
                margin-bottom: 0;
                flex: 1 1 auto;
            }
            .add-title {
                margin: 0 0 6px;
                font-size: 1.05rem;
                color: #1e3a8a;
            }
            .add-help {
                margin: 0 0 14px;
                font-size: 0.9rem;
                color: #334155;
            }
            .add-grid {
                display: grid;
                grid-template-columns: repeat(2, minmax(220px, 1fr));
                gap: 10px 14px;
            }
            @media (max-width: 700px) {
                .add-grid {
                    grid-template-columns: 1fr;
                }
                .add-section {
                    max-width: 100%;
                }
            }
            label {
                display: block;
                font-size: 0.92rem;
                margin-bottom: 4px;
            }
            input[type="text"] {
                width: 100%;
                max-width: 420px;
                padding: 8px;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                box-sizing: border-box;
            }
            .per-page-form input.per-page-input {
                width: 80px;
                max-width: 80px;
            }
            button {
                background: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 12px;
                cursor: pointer;
            }
            .small-btn {
                padding: 6px 10px;
                font-size: 0.85rem;
                text-decoration: none;
                display: inline-block;
                border-radius: 6px;
                background: #2563eb;
                color: white;
            }
            .add-submit {
                margin-top: 6px;
                font-size: 0.9rem;
            }
            .per-page-submit {
                padding: 8px 12px;
                font-size: 0.9rem;
                border: none;
            }
            .delete-btn {
                background: #dc2626;
            }
            .inline-form {
                display: inline;
            }
            .toolbar {
                display: flex;
                justify-content: space-between;
                align-items: flex-start;
                gap: 12px;
                margin: 2px 0 0;
            }
            .toolbar-left {
                display: flex;
                flex-direction: column;
                gap: 4px;
            }
            .toolbar-right {
                display: flex;
                flex-direction: column;
                align-items: flex-end;
                gap: 4px;
            }
            .per-page-form {
                display: flex;
                flex-direction: column;
                align-items: flex-end;
                gap: 4px;
            }
            .per-page-controls {
                display: flex;
                align-items: center;
                gap: 8px;
            }
            .per-page-form label {
                margin-bottom: 0;
            }
            .per-page-input {
                width: 80px;
                max-width: 80px;
                display: inline-block;
            }
            .meta {
                color: #4b5563;
                font-size: 0.92rem;
            }
            .filter-section {
                border: 2px solid #bfdbfe;
                background: #eff6ff;
                border-radius: 10px;
                padding: 16px;
                margin-bottom: 24px;
            }
            .filter-grid {
                display: grid;
                grid-template-columns: repeat(4, minmax(170px, 1fr));
                gap: 10px 12px;
            }
            .filter-actions {
                margin-top: 8px;
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 12px;
                flex-wrap: wrap;
            }
            .filter-main-actions,
            .filter-export-actions {
                display: flex;
                align-items: center;
                gap: 8px;
            }
            .filter-export-actions {
                margin-left: auto;
            }
            .filter-actions button,
            .filter-actions .small-btn {
                height: 38px;
                padding: 0 12px;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                font-size: 0.9rem;
            }
            .pdf-btn {
                background: #d32f2f;
            }
            .pdf-btn:hover {
                background: #b71c1c;
            }
            .excel-btn {
                background: #217346;
            }
            .excel-btn:hover {
                background: #185c37;
            }
            .reset-btn {
                background: #4b5563;
            }
            input[type="number"],
            select {
                width: 100%;
                max-width: 420px;
                padding: 8px;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                box-sizing: border-box;
            }
            @media (max-width: 980px) {
                .filter-grid {
                    grid-template-columns: repeat(2, minmax(160px, 1fr));
                }
            }
            @media (max-width: 620px) {
                .filter-grid {
                    grid-template-columns: 1fr;
                }
            }
            .pager {
                display: flex;
                align-items: center;
                gap: 8px;
                margin-top: 12px;
            }
            .pager a {
                text-decoration: none;
                color: #1d4ed8;
            }
        </style>

        <div class="container">
            <div class="top-layout">
                <div class="brand-panel">
                    <div class="page-head">
                        <div>
                            <h1>Autolijst</h1>
                            <div class="version">Versie: {{ app_version }}</div>
                        </div>
                        {% if header_image_url %}
                            <img src="{{ header_image_url }}" alt="Auto" class="brand-image">
                        {% endif %}
                    </div>
                </div>

                <div class="add-section">
                    <h3 class="add-title">Nieuwe auto invoeren</h3>
                    <p class="add-help">Vul de velden in en klik op Toevoegen om direct een nieuwe rij op te slaan.</p>
                    <form method="post" action="{{ url_for('add_row') }}">
                        <input type="hidden" name="page" value="{{ page }}">
                        <input type="hidden" name="per_page" value="{{ per_page }}">
                        <input type="hidden" name="sort_by" value="{{ sort_by }}">
                        <input type="hidden" name="sort_dir" value="{{ sort_dir }}">
                        {% for key, value in filter_values.items() %}
                            <input type="hidden" name="{{ key }}" value="{{ value }}">
                        {% endfor %}
                        <div class="add-grid">
                            {% for column in editable_columns %}
                                <div class="field">
                                    <label>{{ column }}</label>
                                    <input type="text" name="{{ column }}">
                                </div>
                            {% endfor %}
                        </div>
                        <button type="submit" class="add-submit">Toevoegen</button>
                    </form>
                </div>
            </div>

            <div class="filter-section">
                <h3 class="add-title">Filter</h3>
                <form method="get" action="{{ url_for('home') }}">
                    <input type="hidden" name="page" value="1">
                    <input type="hidden" name="per_page" value="{{ per_page }}">
                    <input type="hidden" name="sort_by" value="{{ sort_by }}">
                    <input type="hidden" name="sort_dir" value="{{ sort_dir }}">
                    <div class="filter-grid">
                        <div class="field">
                            <label for="filter_merk">Merk bevat</label>
                            <input id="filter_merk" type="text" name="filter_merk" value="{{ filter_values['filter_merk'] }}">
                        </div>
                        <div class="field">
                            <label for="filter_type">Type bevat</label>
                            <input id="filter_type" type="text" name="filter_type" value="{{ filter_values['filter_type'] }}">
                        </div>
                        <div class="field">
                            <label for="filter_bouwjaar_min">Bouwjaar van</label>
                            <input id="filter_bouwjaar_min" type="number" name="filter_bouwjaar_min" value="{{ filter_values['filter_bouwjaar_min'] }}">
                        </div>
                        <div class="field">
                            <label for="filter_bouwjaar_max">Bouwjaar t/m</label>
                            <input id="filter_bouwjaar_max" type="number" name="filter_bouwjaar_max" value="{{ filter_values['filter_bouwjaar_max'] }}">
                        </div>
                        <div class="field">
                            <label for="filter_prijs_min">Prijs van</label>
                            <input id="filter_prijs_min" type="text" name="filter_prijs_min" value="{{ filter_values['filter_prijs_min'] }}">
                        </div>
                        <div class="field">
                            <label for="filter_prijs_max">Prijs t/m</label>
                            <input id="filter_prijs_max" type="text" name="filter_prijs_max" value="{{ filter_values['filter_prijs_max'] }}">
                        </div>
                        <div class="field">
                            <label for="filter_categorie">Categorie</label>
                            <select id="filter_categorie" name="filter_categorie">
                                <option value="" {% if not filter_values['filter_categorie'] %}selected{% endif %}>Alle</option>
                                <option value="Klassieker" {% if filter_values['filter_categorie'] == 'Klassieker' %}selected{% endif %}>Klassieker</option>
                                <option value="Youngtimer" {% if filter_values['filter_categorie'] == 'Youngtimer' %}selected{% endif %}>Youngtimer</option>
                            </select>
                        </div>
                    </div>
                    <div class="filter-actions">
                        <div class="filter-main-actions">
                            <button type="submit">Filter</button>
                            <a href="{{ url_for('home', per_page=per_page, sort_by=sort_by, sort_dir=sort_dir) }}" class="small-btn reset-btn">Reset</a>
                        </div>
                        <div class="filter-export-actions">
                            <a href="{{ pdf_export_url }}" class="small-btn pdf-btn">PDF van selectie</a>
                            <a href="{{ excel_export_url }}" class="small-btn excel-btn">Excel van selectie</a>
                        </div>
                    </div>
                </form>
            </div>

            <div class="toolbar">
                <div class="toolbar-left">
                    <h2>Overzicht</h2>
                    <div class="meta">Totaal rijen: <strong>{{ total_rows }}</strong></div>
                </div>
                <div class="toolbar-right">
                    <form method="get" action="{{ url_for('home') }}" class="per-page-form">
                        <label for="per_page">Rijen per pagina</label>
                        <div class="per-page-controls">
                            <input id="per_page" type="text" name="per_page" value="{{ per_page }}" class="per-page-input">
                            <input type="hidden" name="page" value="1">
                            <input type="hidden" name="sort_by" value="{{ sort_by }}">
                            <input type="hidden" name="sort_dir" value="{{ sort_dir }}">
                            {% for key, value in filter_values.items() %}
                                <input type="hidden" name="{{ key }}" value="{{ value }}">
                            {% endfor %}
                            <button type="submit" class="per-page-submit">Toon</button>
                        </div>
                    </form>
                </div>
            </div>

            <table class="data-table">
                <thead>
                    <tr>
                        <th><a href="{{ sort_urls['id'] }}" class="sort-link">{{ sort_labels['id'] }}</a></th>
                        {% for column in columns %}
                            <th><a href="{{ sort_urls[column] }}" class="sort-link">{{ sort_labels[column] }}</a></th>
                        {% endfor %}
                        <th>Acties</th>
                    </tr>
                </thead>
                <tbody>
                    {% if records %}
                        {% for record in records %}
                            <tr>
                                <td>{{ record.display_id }}</td>
                                {% for column in columns %}
                                    <td>{{ record.data[column] }}</td>
                                {% endfor %}
                                <td>
                                    <a href="{{ url_for('edit_row', row_index=record.index, page=page, per_page=per_page, sort_by=sort_by, sort_dir=sort_dir, filter_merk=filter_values['filter_merk'], filter_type=filter_values['filter_type'], filter_bouwjaar_min=filter_values['filter_bouwjaar_min'], filter_bouwjaar_max=filter_values['filter_bouwjaar_max'], filter_prijs_min=filter_values['filter_prijs_min'], filter_prijs_max=filter_values['filter_prijs_max'], filter_categorie=filter_values['filter_categorie']) }}" class="small-btn">Bewerk</a>
                                    <form method="post" action="{{ url_for('delete_row', row_index=record.index) }}" class="inline-form">
                                        <input type="hidden" name="page" value="{{ page }}">
                                        <input type="hidden" name="per_page" value="{{ per_page }}">
                                        <input type="hidden" name="sort_by" value="{{ sort_by }}">
                                        <input type="hidden" name="sort_dir" value="{{ sort_dir }}">
                                        {% for key, value in filter_values.items() %}
                                            <input type="hidden" name="{{ key }}" value="{{ value }}">
                                        {% endfor %}
                                        <button type="submit" class="small-btn delete-btn">Verwijder</button>
                                    </form>
                                </td>
                            </tr>
                        {% endfor %}
                    {% else %}
                        <tr>
                            <td colspan="{{ columns|length + 2 }}">Er zijn nog geen rijen aanwezig.</td>
                        </tr>
                    {% endif %}
                </tbody>
            </table>

            <div class="pager">
                {% if prev_page_url %}
                    <a href="{{ prev_page_url }}">← Vorige</a>
                {% endif %}
                <span>Pagina {{ page }} van {{ total_pages }}</span>
                {% if next_page_url %}
                    <a href="{{ next_page_url }}">Volgende →</a>
                {% endif %}
            </div>
        </div>
        """,
        columns=columns,
        editable_columns=editable_columns,
        records=records,
        total_rows=total_rows,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        app_version=app_version,
        header_image_url=header_image_url,
        filter_values=filter_values,
        sort_by=sort_by,
        sort_dir=sort_dir,
        sort_urls=sort_urls,
        sort_labels=sort_labels,
        prev_page_url=prev_page_url,
        next_page_url=next_page_url,
        pdf_export_url=pdf_export_url,
        excel_export_url=excel_export_url,
    )


@app.route("/version")
def version():
    return jsonify({"version": get_app_version()})


@app.route("/export/pdf")
def export_pdf():
    df = load_dataframe()
    df, _ = ensure_category_column(df)
    df, _ = apply_category_rules(df)

    filter_values = get_filter_values_from_request(request.args)
    filtered_df = build_filtered_dataframe(df, filter_values)
    sort_by, sort_dir = get_sort_values_from_request(request.args)
    filtered_df = apply_sorting(filtered_df, sort_by, sort_dir)
    columns = list(df.columns)

    try:
        pagesizes = importlib.import_module("reportlab.lib.pagesizes")
        pdf_canvas = importlib.import_module("reportlab.pdfgen.canvas")
        pdf_colors = importlib.import_module("reportlab.lib.colors")
        a4_size = pagesizes.A4
        canvas_class = pdf_canvas.Canvas
    except Exception:
        return "PDF-export vereist reportlab. Installeer dependencies met pip install -r requirements.txt.", 500

    buffer = BytesIO()
    pdf = canvas_class(buffer, pagesize=a4_size)

    page_width, page_height = a4_size
    left_margin = 30
    right_margin = 30
    top_margin = 32
    bottom_margin = 28
    usable_width = page_width - left_margin - right_margin

    header_bg = pdf_colors.Color(0.90, 0.95, 1)
    table_header_bg = pdf_colors.Color(0.16, 0.39, 0.92)
    table_header_text = pdf_colors.white
    row_alt_bg = pdf_colors.Color(0.97, 0.98, 1)
    border_color = pdf_colors.Color(0.82, 0.84, 0.88)
    normal_text = pdf_colors.Color(0.12, 0.16, 0.22)

    table_columns = [column for column in columns if column in ["Merk", "Type", "Bouwjaar", "Prijs", "Categorie"]]
    column_weights = {
        "Merk": 0.19,
        "Type": 0.28,
        "Bouwjaar": 0.15,
        "Prijs": 0.17,
        "Categorie": 0.21,
    }
    widths = [usable_width * column_weights[column] for column in table_columns]

    def truncate_for_width(text: str, width: float) -> str:
        max_chars = max(6, int(width / 5.2))
        if len(text) <= max_chars:
            return text
        return text[: max_chars - 1] + "…"

    def draw_table_header(y_pos: float) -> float:
        pdf.setStrokeColor(border_color)
        x_pos = left_margin
        header_height = 20
        for index, column in enumerate(table_columns):
            cell_width = widths[index]
            pdf.setFillColor(table_header_bg)
            pdf.rect(x_pos, y_pos - header_height, cell_width, header_height, stroke=1, fill=1)
            pdf.setFillColor(table_header_text)
            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(x_pos + 5, y_pos - 13, column)
            x_pos += cell_width
        return y_pos - header_height

    now_text = datetime.now().strftime("%d-%m-%Y %H:%M")
    active_filters = [f"{key.replace('filter_', '')}: {value}" for key, value in filter_values.items() if value]
    filter_text = " | ".join(active_filters) if active_filters else "Geen"

    title_box_height = 76
    y = page_height - top_margin

    pdf.setStrokeColor(border_color)
    pdf.setFillColor(header_bg)
    pdf.roundRect(left_margin, y - title_box_height, usable_width, title_box_height, 8, stroke=1, fill=1)

    pdf.setFillColor(normal_text)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(left_margin + 12, y - 24, "Autolijst - Gefilterde selectie")
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left_margin + 12, y - 40, f"Exportmoment: {now_text}")
    pdf.drawString(left_margin + 12, y - 54, f"Aantal rijen: {len(filtered_df)}")
    pdf.drawString(left_margin + 12, y - 68, truncate_for_width(f"Filters: {filter_text}", usable_width - 24))

    y -= title_box_height + 16
    y = draw_table_header(y)

    if filtered_df.empty:
        row_height = 18
        pdf.setFillColor(pdf_colors.white)
        pdf.rect(left_margin, y - row_height, usable_width, row_height, stroke=1, fill=1)
        pdf.setFillColor(normal_text)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(left_margin + 6, y - 12, "Geen rijen gevonden voor de gekozen filters.")
    else:
        row_height = 17
        row_counter = 0
        for row_index in filtered_df.index:
            if y - row_height < bottom_margin:
                pdf.showPage()
                y = page_height - top_margin
                y = draw_table_header(y)

            row_counter += 1
            x_pos = left_margin
            for col_index, column in enumerate(table_columns):
                cell_width = widths[col_index]

                if row_counter % 2 == 0:
                    pdf.setFillColor(row_alt_bg)
                else:
                    pdf.setFillColor(pdf_colors.white)

                pdf.setStrokeColor(border_color)
                pdf.rect(x_pos, y - row_height, cell_width, row_height, stroke=1, fill=1)

                cell_value = to_display_value(
                    filtered_df.at[row_index, column],
                    column_name=column,
                    format_price=True,
                )
                text_value = truncate_for_width(str(cell_value), cell_width - 10)

                pdf.setFillColor(normal_text)
                pdf.setFont("Helvetica", 8.5)
                pdf.drawString(x_pos + 5, y - 12, text_value)
                x_pos += cell_width

            y -= row_height

    pdf.save()
    buffer.seek(0)

    file_name = f"autolijst_selectie_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=file_name)


@app.route("/export/excel")
def export_excel():
    df = load_dataframe()
    df, _ = ensure_category_column(df)
    df, _ = apply_category_rules(df)

    filter_values = get_filter_values_from_request(request.args)
    filtered_df = build_filtered_dataframe(df, filter_values)
    sort_by, sort_dir = get_sort_values_from_request(request.args)
    filtered_df = apply_sorting(filtered_df, sort_by, sort_dir)
    columns = [column for column in ["Merk", "Type", "Bouwjaar", "Prijs", "Categorie"] if column in df.columns]

    if not columns:
        columns = list(df.columns)

    export_rows = []
    for row_index in filtered_df.index:
        export_rows.append(
            {
                column: to_display_value(
                    filtered_df.at[row_index, column],
                    column_name=column,
                    format_price=True,
                )
                for column in columns
            }
        )

    export_df = pd.DataFrame(export_rows, columns=columns)
    if export_df.empty:
        empty_row = {column: "" for column in columns}
        if columns:
            empty_row[columns[0]] = "Geen rijen gevonden voor de gekozen filters."
        export_df = pd.DataFrame([empty_row], columns=columns)

    active_filters = [f"{key.replace('filter_', '')}: {value}" for key, value in filter_values.items() if value]
    filter_text = " | ".join(active_filters) if active_filters else "Geen"

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Selectie", startrow=5)

        worksheet = writer.sheets["Selectie"]
        workbook = writer.book

        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo

        title_fill = PatternFill("solid", fgColor="E6F0FF")
        title_font = Font(bold=True, size=14)
        meta_font = Font(size=10)

        worksheet["A1"] = "Autolijst - Gefilterde selectie"
        worksheet["A2"] = f"Exportmoment: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        worksheet["A3"] = f"Aantal rijen: {len(filtered_df)}"
        worksheet["A4"] = f"Filters: {filter_text}"

        for row_number in [1, 2, 3, 4]:
            cell = worksheet[f"A{row_number}"]
            cell.fill = title_fill
            cell.font = title_font if row_number == 1 else meta_font

        max_header_col = max(1, len(columns))
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_header_col)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_header_col)
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_header_col)
        worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=max_header_col)

        header_row = 6
        last_row = header_row + len(export_df)

        from openpyxl.utils import get_column_letter

        column_widths = {
            "Merk": 20,
            "Type": 30,
            "Bouwjaar": 14,
            "Prijs": 16,
            "Categorie": 18,
        }

        for index, column in enumerate(columns, start=1):
            column_letter = get_column_letter(index)
            worksheet.column_dimensions[column_letter].width = column_widths.get(column, 18)

        table_ref = f"A{header_row}:{get_column_letter(len(columns))}{last_row}"
        table = Table(displayName="AutolijstSelectie", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        worksheet.freeze_panes = "A7"

    buffer.seek(0)
    file_name = f"autolijst_selectie_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=file_name,
    )


@app.route("/add", methods=["POST"])
def add_row():
    df = load_dataframe()
    if len(df.columns) == 0:
        return redirect(url_for("home"))

    df, _ = ensure_category_column(df)

    new_row = {column: request.form.get(column, "") for column in default_columns}
    if "id" in df.columns:
        new_row["id"] = pd.NA

    df.loc[len(df)] = new_row
    df, _ = apply_category_rules(df)
    try:
        save_dataframe(df)
    except PermissionError:
        return render_file_locked_message("rij toevoegen")

    filter_values = get_filter_values_from_request(request.form)
    non_empty_filter_params = get_non_empty_filter_params(filter_values)
    sort_by, sort_dir = get_sort_values_from_request(request.form)

    return redirect(
        url_for(
            "home",
            page=parse_positive_int(request.form.get("page"), 1),
            per_page=parse_positive_int(request.form.get("per_page"), 50),
            sort_by=sort_by,
            sort_dir=sort_dir,
            **non_empty_filter_params,
        )
    )


@app.route("/edit/<int:row_index>", methods=["GET", "POST"])
def edit_row(row_index: int):
    df = load_dataframe()
    df, _ = ensure_category_column(df)
    df, _ = apply_category_rules(df)
    if row_index < 0 or row_index >= len(df):
        return redirect(url_for("home"))

    page = parse_positive_int(request.values.get("page"), 1)
    per_page = parse_positive_int(request.values.get("per_page"), 50)
    filter_values = get_filter_values_from_request(request.values)
    sort_by, sort_dir = get_sort_values_from_request(request.values)
    editable_columns = [column for column in df.columns if column not in ["id", "Categorie"]]

    if request.method == "POST":
        for column in editable_columns:
            df.at[row_index, column] = request.form.get(column, "")

        df, _ = apply_category_rules(df)

        try:
            save_dataframe(df)
        except PermissionError:
            return render_file_locked_message("rij bewerken")

        non_empty_filter_params = get_non_empty_filter_params(filter_values)
        return redirect(
            url_for(
                "home",
                page=page,
                per_page=per_page,
                sort_by=sort_by,
                sort_dir=sort_dir,
                **non_empty_filter_params,
            )
        )

    row_data = {
        column: to_display_value(df.iloc[row_index][column])
        for column in df.columns
    }

    return render_template_string(
        """
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 0;
                background: #f6f8fb;
                color: #1f2937;
            }
            .container {
                max-width: 800px;
                margin: 28px auto;
                background: white;
                border: 1px solid #e5e7eb;
                border-radius: 10px;
                padding: 24px;
            }
            .field {
                margin-bottom: 10px;
            }
            label {
                display: block;
                font-size: 0.92rem;
                margin-bottom: 4px;
            }
            input[type="text"] {
                width: 100%;
                max-width: 520px;
                padding: 8px;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                box-sizing: border-box;
            }
            button,
            a {
                background: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 12px;
                text-decoration: none;
                cursor: pointer;
                display: inline-block;
                font-size: 0.9rem;
            }
            .actions {
                margin-top: 14px;
                display: flex;
                gap: 8px;
            }
            .back-link {
                background: #4b5563;
            }
        </style>

        <div class="container">
            <h1>Rij {{ row_number }} bewerken</h1>
            <form method="post" action="{{ url_for('edit_row', row_index=row_index) }}">
                <input type="hidden" name="page" value="{{ page }}">
                <input type="hidden" name="per_page" value="{{ per_page }}">
                <input type="hidden" name="sort_by" value="{{ sort_by }}">
                <input type="hidden" name="sort_dir" value="{{ sort_dir }}">
                {% for key, value in filter_values.items() %}
                    <input type="hidden" name="{{ key }}" value="{{ value }}">
                {% endfor %}
                {% for column in editable_columns %}
                    <div class="field">
                        <label>{{ column }}</label>
                        <input type="text" name="{{ column }}" value="{{ row_data[column] }}">
                    </div>
                {% endfor %}
                <div class="field">
                    <label>Categorie (automatisch)</label>
                    <input type="text" value="{{ row_data['Categorie'] }}" readonly>
                </div>
                <div class="actions">
                    <button type="submit">Opslaan</button>
                    <a href="{{ url_for('home', page=page, per_page=per_page, sort_by=sort_by, sort_dir=sort_dir, filter_merk=filter_values['filter_merk'], filter_type=filter_values['filter_type'], filter_bouwjaar_min=filter_values['filter_bouwjaar_min'], filter_bouwjaar_max=filter_values['filter_bouwjaar_max'], filter_prijs_min=filter_values['filter_prijs_min'], filter_prijs_max=filter_values['filter_prijs_max'], filter_categorie=filter_values['filter_categorie']) }}" class="back-link">Terug</a>
                </div>
            </form>
        </div>
        """,
        editable_columns=editable_columns,
        row_data=row_data,
        row_number=row_index + 1,
        row_index=row_index,
        page=page,
        per_page=per_page,
        sort_by=sort_by,
        sort_dir=sort_dir,
        filter_values=filter_values,
    )


@app.route("/delete/<int:row_index>", methods=["POST"])
def delete_row(row_index: int):
    df = load_dataframe()
    df, _ = ensure_category_column(df)
    df, _ = apply_category_rules(df)
    if row_index < 0 or row_index >= len(df):
        return redirect(url_for("home"))

    df = df.drop(index=row_index).reset_index(drop=True)
    try:
        save_dataframe(df)
    except PermissionError:
        return render_file_locked_message("rij verwijderen")

    filter_values = get_filter_values_from_request(request.form)
    non_empty_filter_params = get_non_empty_filter_params(filter_values)
    sort_by, sort_dir = get_sort_values_from_request(request.form)

    return redirect(
        url_for(
            "home",
            page=parse_positive_int(request.form.get("page"), 1),
            per_page=parse_positive_int(request.form.get("per_page"), 50),
            sort_by=sort_by,
            sort_dir=sort_dir,
            **non_empty_filter_params,
        )
    )

if __name__ == "__main__":
    host = os.getenv("HOST", "0.0.0.0")
    port = parse_positive_int(os.getenv("PORT"), 5000)
    debug = os.getenv("FLASK_DEBUG", "1") == "1"
    app.run(host=host, port=port, debug=debug)